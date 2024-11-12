import re  # Importa il modulo regex
import pandas as pd
import os
import cx_Oracle
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle
import logging
from datetime import datetime

# Configura il logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class OracleDB:
    def __init__(self):
        self.username = 'C4SF0_APPP'
        self.password = 'prjcrt18'
        self.hostname = 'scaninfr010prd.sede.corp.sanpaoloimi.com'
        self.port = '1521'
        self.service_name = 'SC4SF0ONP'
        self.dsn = cx_Oracle.makedsn(self.hostname, self.port, service_name=self.service_name)
        self.connection = None
        self.logger = logger
    
    def connect(self):
        try:
            self.connection = cx_Oracle.connect(self.username, self.password, self.dsn)
            self.logger.debug("Connessione db stabilita")
        except cx_Oracle.DatabaseError as e:
            self.logger.error(f"Errore di connessione a DB: {e}")
            self.connection = None
    
    def disconnect(self):
        if self.connection:
            self.connection.close()
            self.logger.debug("Connessione db chiusa")
        else:
            self.logger.debug("Non sono presenti connessioni a DB aperte")

    def query_to_dataframe(self, query, params=None):
        if self.connection:
            cursor = self.connection.cursor()
            try:
                cursor.execute(query, params or {})
            
                # Estrazione delle colonne
                columns = [col[0] for col in cursor.description]
            
                # Estrazione delle righe direttamente dal cursore
                result = []
                for row in cursor:
                    result.append(row)
            
                # Creazione del dataframe
                df = pd.DataFrame(result, columns=columns)
                return df
            except cx_Oracle.DatabaseError as e:
                self.logger.debug(f"Errore nel lancio della query DB: {e}")
                return None
            finally:
                cursor.close()
        else:
            self.logger.debug("Non sono presenti connessioni aperte. Connettiti prima al db")
            return None


    
#    def query_to_dataframe(self, query, params=None):
#        if self.connection:
#            cursor = self.connection.cursor()
#            try:
#                cursor.execute(query, params or {})
#                columns = [col[0] for col in cursor.description]  # Colonne della tabella
#                result = cursor.fetchall()
#                df = pd.DataFrame(result, columns=columns)
#                return df
#            except cx_Oracle.DatabaseError as e:
#                self.logger.debug(f"Errore nel lancio della query DB: {e}")
#                return None
#            finally:
#                cursor.close()
#        else:
#            self.logger.debug("Non sono presenti connessioni aperte. Connettiti prima al db")
#            return None

def format_sheet(ws):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.border = thick_border

    # Ottimizzazione larghezza colonne senza ciclo su celle vuote
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Formattazione delle date e dei numeri
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')  # Formato numerico con due decimali

    header = [cell.value for cell in ws[1]]  # Ottieni i nomi delle intestazioni
    date_columns = [header.index(col) + 1 for col in ['Trade Date', 'Start Date', 'End Date'] if col in header]  # Trova le posizioni delle colonne delle date


    number_columns = ["Client Code"]
#    number_columns = [
#        "Client Code", "Agent Code", "Trade Xref 1", "Trade Xref Type 1",
#        "Trade Xref 2", "Trade Xref Type 2", "Trade Xref 3", "Trade Xref Type 3",
#        "Trade Xref 4", "Trade Xref Type 4", "Trade Xref 5", "Trade Xref Type 5",
#        "Trade Xref 6", "Trade Xref Type 6"
#    ]
    
    # Trova le posizioni delle colonne numeriche
    number_columns_indices = [header.index(col) + 1 for col in number_columns if col in header]

    for row in ws.iter_rows(min_row=2):  # Salta l'intestazione
        for cell in row:
            logger.info(f"Indice della colonna corrente: {cell.column}")
            if cell.column in date_columns:  # Controlla se la colonna è una delle colonne delle date
                if isinstance(cell.value, (datetime, str)):  # Controlla se il valore è una data
                    cell.style = date_style
                else:
                    print("non è di tipo data")
            elif cell.column in number_columns_indices:  # Controlla se la colonna è una delle colonne numeriche
                # Verifica se il valore della cella contiene solo i caratteri consentiti
                if isinstance(cell.value, (int, float, str)):  # Controlla se il valore è un numero o una stringa
                    if re.fullmatch(r'^[0123456789.,]+$', str(cell.value)): 
                        logger.info(f"Cell coordinate: {cell.coordinate}, Column: {cell.column}, Value type: {type(cell.value).__name__}")  # Stampa la cella, la colonna e il tipo di valore
                        # ws[cell.coordinate].number_format  # Applica il formato della cella di riferimento
                        cell.value = int(cell.value)
                        logger.info(f"Cell coordinate: {cell.coordinate}, Column: {cell.column}, Valore nuovo: {type(cell.value).__name__}")  # Stampa la cella, la colonna e il tipo di valore
                        

def main():
    data_corrente = datetime.now().strftime("%d%m%y")
    excel_path = os.path.dirname(os.path.abspath(__file__))
    db = OracleDB()
    db.connect()

    #sql_folder = "./sql"
    sql_folder = os.path.join(excel_path, "sql")
    cib_file, isp_file = None, None

    for file in os.listdir(sql_folder):
        if file.lower().startswith('cib') and file.endswith('.sql'):
            cib_file = file
        elif file.lower().startswith('isp') and file.endswith('.sql'):
            isp_file = file

    if cib_file:
        with open(os.path.join(sql_folder, cib_file), 'r') as file:
            cib_query = file.read()
        CIB_ds = db.query_to_dataframe(cib_query)
        logger.info("Dataset CIB creato con successo.")
    else:
        logger.info("Nessun file CIB trovato nella cartella sql.")

    if isp_file:
        with open(os.path.join(sql_folder, isp_file), 'r') as file:
            isp_query = file.read()
        ISP_ds = db.query_to_dataframe(isp_query)
        logger.info("Dataset ISP creato con successo.")
    else:
        logger.info("Nessun file ISP trovato nella cartella sql.")

    db.disconnect()

    # Creiamo un nuovo Workbook per CIB
    wb_cib = Workbook()
    ws_cib = wb_cib.active
    ws_cib.title = "CIB"
    for r in dataframe_to_rows(CIB_ds, index=False, header=True):
        ws_cib.append(r)
    format_sheet(ws_cib)  # Applica la formattazione
    # Salvataggio del file Excel per CIB
    excel_filename_cib = os.path.join(excel_path, f"Estrazione_TradeEnquiry_{data_corrente}_CIB.xlsx")
    wb_cib.save(excel_filename_cib)
    logger.info(f"File Excel CIB salvato in: {excel_filename_cib}")

    # Creiamo un nuovo Workbook per ISP
    wb_isp = Workbook()
    ws_isp = wb_isp.active
    ws_isp.title = "ISP"
    for r in dataframe_to_rows(ISP_ds, index=False, header=True):
        ws_isp.append(r)
    format_sheet(ws_isp)  # Applica la formattazione
    # Salvataggio del file Excel per ISP
    excel_filename_isp = os.path.join(excel_path, f"Estrazione_TradeEnquiry_{data_corrente}_ISP.xlsx")
    wb_isp.save(excel_filename_isp)
    logger.info(f"File Excel ISP salvato in: {excel_filename_isp}")

    if 'CIB_ds' in locals() and CIB_ds is not None:
        logger.info("\nPrime righe del dataset CIB:")
        logger.info(CIB_ds.head())
        logger.info(f"Numero totale di righe nel dataset CIB: {len(CIB_ds)}")

    if 'ISP_ds' in locals() and ISP_ds is not None:
        logger.info("\nPrime righe del dataset ISP:")
        logger.info(ISP_ds.head())

if __name__ == "__main__":
    main()

